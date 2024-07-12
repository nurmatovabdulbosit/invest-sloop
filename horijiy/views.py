from django.shortcuts import render
from .models import Horijiy
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.contrib import admin
from import_export.admin import ImportMixin
from .resources import HorijiyResource


class HorijiyAdmin(ImportMixin, admin.ModelAdmin):
    resource_class = HorijiyResource

    def get_import_resource_kwargs(self, request, *args, **kwargs):
        return {'user': request.user}

def export_to_excel(request):
    # Horijiy obyektlarini olish
    horijiy_list = Horijiy.objects.all()

    # Excel faylni yaratish
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'manzilli'

    # Ma'lumotlar ustuni joylash
    headers = [
        'T/R', 'Лойиҳа ташаббускорлари ва лойиҳа номи', 'Ҳудудлар номи / бариктирилган маъсуллар', 
        'Туман ва шаҳарлар номи','Лойиҳа қуввати','Амалга ошириш муддати', 'Давлат номи', 
        'Тармоқ/соҳа йўналиши', 'Лойиҳа умумий қиймати','Лойиҳани бошланган санаси', 'PDF', 'Excel'
    ]
    bold_font = Font(bold=True)
    worksheet.append(headers)
    for cell in worksheet["1:1"]:
        cell.font = bold_font

    # Ma'lumotlarni Excelga yozish
    for index, horijiy in enumerate(horijiy_list, start=1):
        row = [
            index,
            horijiy.title,
            horijiy.masul_vakil,
            horijiy.get_tuman_display(),
            horijiy.loyiha_quvvati,
            horijiy.tugatish_sanasi.strftime('%d.%m.%Y') if horijiy.tugatish_sanasi else '',
            horijiy.davlat,
            horijiy.get_soha_yonalishi_display(),
            horijiy.loyiha_qiymati,
            horijiy.create_at.strftime('%d.%m.%Y') if horijiy.create_at else '',
        ]

        # URL manzillarini yaratish
        pdf_url = horijiy.pdf_file.url if horijiy.pdf_file else ''
        excel_url = horijiy.excel_file.url if horijiy.excel_file else ''

        # PDF va Excel havolalarini qo'shish
        row.append(pdf_url)
        row.append(excel_url)
        
        worksheet.append(row)

        # PDF va Excel giperhavolalarni qo'shish
        pdf_cell = worksheet.cell(row=index+1, column=len(headers)-1)
        if pdf_url:
            pdf_cell.hyperlink = pdf_url
            pdf_cell.value = "PDF"
            pdf_cell.style = "Hyperlink"

        excel_cell = worksheet.cell(row=index+1, column=len(headers))
        if excel_url:
            excel_cell.hyperlink = excel_url
            excel_cell.value = "Excel"
            excel_cell.style = "Hyperlink"

    # HttpResponse orqali Excel faylni yuborish
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="horijiy_data.xlsx"'
    workbook.save(response)

    return response

def index(request):
    horijiy_list = Horijiy.objects.all()
    
    # Status bo'yicha filtratsiya
    status = request.GET.get('status')
    if status:
        horijiy_list = horijiy_list.filter(status=status)
    
    # Soha yonalishi bo'yicha filtratsiya
    soha_yonalishi = request.GET.get('soha_yonalishi')
    if soha_yonalishi:
        horijiy_list = horijiy_list.filter(soha_yonalishi=soha_yonalishi)

    # Jami qiymat hisoblanishi
    total_sum = horijiy_list.aggregate(total_sum=Sum('loyiha_qiymati'))['total_sum']

    context = {
        'horijiy_list': horijiy_list,
        'total_sum': total_sum,
    }
    return render(request, 'index.html', context)
def index(request):
    active_count = Horijiy.objects.filter(status='active').count()
    deactive_count = Horijiy.objects.filter(status='deactive').count()
    expired_count = Horijiy.objects.filter(status='expired').count()
    tuman = request.GET.get('tuman', '')
    horijiy_list = Horijiy.objects.all()
    if tuman:
        horijiy_list = horijiy_list.filter(tuman=tuman)

    total_sum = horijiy_list.aggregate(total_sum=Sum('loyiha_qiymati'))['total_sum']

    context = {
        'active_count': active_count,
        'deactive_count': deactive_count,
        'expired_count': expired_count,
        'horijiy_list': horijiy_list,
        'total_sum': total_sum,
        'selected_tuman': tuman,  # Pass the selected value to the template
        'TUMANLAR': Horijiy.TUMANLAR,
    }

    return render(request, 'index.html', context)